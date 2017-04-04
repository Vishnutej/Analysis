from openpyxl import Workbook
from openpyxl import load_workbook
from sets import Set
def sum():
	print "Hello"
	s=Set()
	wb = load_workbook(filename = 'C:\Users\Vishnutej\Desktop\inhersight.xlsx')
	#sheet_ranges = wb['Sheet1'] 
	sheet = wb.get_sheet_by_name('Sheet1')
	#for row in sheet_ranges.iter_rows(): 
	#		for cell in row:
    #    			print cell.value
	lists = []
	for i in range(1,25):
		lists.append(0)
	
	#print lists
	count=0
	for row in range(2,sheet.max_row):
		bla=str(row)
		print "here"
		print bla
		company = sheet['D' + bla].value
		status  = sheet['E' + bla].value
		if status != None and company !=None:
			s.add(company)
			if count<len(s):
				print status
				lists[status]=lists[status]+1
				count=len(s)
		#print status
	#print(sheet_ranges['A1'].value)
	print 'set'
	print lists
if __name__ == '__main__':
	sum()