from openpyxl import Workbook
from openpyxl import load_workbook
from sets import Set

def sum():
	print "Hello"
	s=Set()
	wb = load_workbook(filename = 'C:\Users\Vishnutej\Desktop\inhersight.xlsx')#Change path as required
	#sheet_ranges = wb['Sheet1'] 
	sheet = wb.get_sheet_by_name('Sheet1')
	for row in range(2,sheet.max_row):
		bla=str(row)
		company = sheet['C'+bla].value
		if company!=None:
			s.add(company)
 	print len(s)
	lists = []
	for i in range(0,len(s)):
		lists.append(0)
	l=[]
	l=list(s)
	l.sort()
	count=0
	print "here"
	for entry in l:
		flag=0
		for row in range(2,sheet.max_row):
			bla=str(row)
			comp=sheet['C'+bla].value
			par=sheet['Q'+bla].value
			if par == None:
				flag=0
			if entry==comp and par !=None:
				lists[count]=lists[count]+1
				flag=1
			else:
				if flag==1 and entry!=comp:
					break
		count=count+1
		if count%100==0:
			#break
			print count#You'll know where the loop is running
	print "loop done -- look for results"
	count=0;
	print len(l)
	for entry in l:
		#if lists[count] == 1:#Change the number of required reviews here
		if lists[count] >=2:
			print str(entry)+" "+str(lists[count])
		count=count+1
		#print count
	print "end"
	#print lists

if __name__ == '__main__':
	sum()